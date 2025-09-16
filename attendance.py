import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import pandas as pd
import os
import glob
from datetime import datetime

class Attendance:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1366x768+0+0")
        self.root.title("Attendance Detail")

        # Create a style for ttk widgets
        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=("Times New Roman", 20), foreground='navy')
        self.style.configure('Button.TButton', font=("times new roman", 15, "bold"), background='navy', foreground='black')
        self.style.map('Button.TButton',
            background=[('active', '#007acc')],
            foreground=[('active', 'blue')]
        )

        # Create a frame to hold the table
        main_frame = ttk.Frame(self.root, relief="ridge")
        main_frame.grid(row=0, column=0, sticky="nsew")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Create a label for "Attendance Detail"
        title_label = ttk.Label(main_frame, text="Attendance Sheet", style='Title.TLabel')
        title_label.pack(side=tk.TOP, padx=10, pady=10)

         # Add a label for subject name at the top
        self.subject_label = ttk.Label(main_frame, text="", style='Title.TLabel')
        self.subject_label.pack(side=tk.TOP, padx=10, pady=10)


        # Create a search frame at the top
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(side=tk.TOP, pady=10)

        self.search_entry = ttk.Entry(search_frame, width=30, font=("Roboto", 12))
        self.search_entry.pack(side=tk.LEFT, padx=10, pady=5)

        self.search_entry.bind("<Return>", lambda event: self.search_data())

        search_button = ttk.Button(search_frame, text="Search", style='Button.TButton', command=self.search_data)
        search_button.pack(side=tk.LEFT)

        y_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL)
        x_scrollbar = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL)

        self.table = ttk.Treeview(main_frame, yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        y_scrollbar.config(command=self.table.yview)
        x_scrollbar.config(command=self.table.xview)
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.table.pack(fill=tk.BOTH, expand=True)

        import_button = ttk.Button(main_frame, text="Import Excel", command=self.import_excel, style='Button.TButton')
        import_button.pack(side=tk.LEFT, padx=(100, 10), pady=10) 

        show_summary_button = ttk.Button(main_frame, text="Show Latest Monthly Summary", command=self.show_monthly_summary, style='Button.TButton')
        show_summary_button.pack(side=tk.LEFT, padx=10, pady=10)

       

    def import_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")], parent=self.root)

        if file_path:
            file_name = os.path.basename(file_path)
            folder_path = "Attendance_sheet"
            full_path = os.path.join(folder_path, file_name)

            wb = load_workbook(file_path)
            sheet = wb.active

            self.table.delete(*self.table.get_children())

            self.table.column("#0", width=1)

            headers = sheet[1]
            self.table["columns"] = headers

            for i, header in enumerate(headers, start=0):
                self.table.heading(i, text=header.value)
                column_width = max(len(str(cell.value)) for cell in sheet[header.column_letter])
                self.table.column(i, width=column_width * 10)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                cleaned_row = [cell if cell is not None and not pd.isna(cell) else 'Absent' for cell in row]
                self.table.insert("", tk.END, values=cleaned_row)

            self.table.xview_moveto(0)
        self.bold_headers()

    def show_monthly_summary(self):
        folder_path = "Attendance_sheet"
        summary_files = glob.glob(os.path.join(folder_path, "attendance_summary_*.xlsx"))

        if summary_files:
            summary_files.sort(key=os.path.getmtime, reverse=True)
            latest_summary_file = summary_files[0]

            df_summary = pd.read_excel(latest_summary_file)
            df_summary = df_summary.fillna('Absent')

            self.table.delete(*self.table.get_children())

            self.table.column("#0", width=1)

            headers = df_summary.columns
            self.table["columns"] = headers

            for i, header in enumerate(headers, start=0):
                self.table.heading(i, text=header)
                column_width = max(df_summary[header].astype(str).apply(len)) + 10
                self.table.column(i, width=column_width * 10)

            for index, row in df_summary.iterrows():
                self.table.insert("", tk.END, values=row.tolist())

            # Update the subject label
            self.subject_label.config(text="Monthly Summary")

        else:
            messagebox.showinfo("Info", "No monthly summary files found in the current directory.")

        self.bold_headers()

    def search_data(self):
        query = self.search_entry.get()
        if query:
            for item in self.table.get_children():
                if query.lower() in str(self.table.item(item)["values"]).lower():
                    self.table.selection_set(item)
                else:
                    self.table.selection_remove(item)

    def bold_headers(self):
        font = ("Roboto", 10, "bold")
        style = ttk.Style()
        style.configure("Treeview.Heading", font=font)

        style.configure("Treeview", rowheight=30, font=("Roboto", 10))
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        style.layout("Treeview.Cell", [('Treeitem.padding', {'sticky': 'nswe', 'children':
            [('Treeitem.border', {'border': '1', 'sticky': 'nswe', 'children': [('Treeitem.text', {'sticky': 'nswe'})]})]})])

    

if __name__ == "__main__":
    root = tk.Tk()
    app = Attendance(root)
    root.mainloop()